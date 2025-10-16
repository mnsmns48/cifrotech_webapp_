from io import BytesIO
from fastapi import Depends, APIRouter, Form
from num2words import num2words
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook
from pydantic import BaseModel, Field
from starlette.requests import Request
from starlette.responses import HTMLResponse, StreamingResponse
from xlsxtpl.writerx import BookWriter

from api_service.schemas.parsing_schemas import ParsingResultOut
from config import BASE_DIR

printer_router = APIRouter(tags=['Service-Bill-Printer'])


class ReceiptForm(BaseModel):
    receipt_date: str = Field(..., alias="receiptDate")
    receipt_number: int = Field(..., alias="receiptNumber")
    receipt_product: str = Field(..., alias="receiptProduct")
    receipt_qty: int = Field(..., alias="receiptQty")
    receipt_price: float = Field(..., alias="receiptPrice")

    @property
    def total(self) -> float:
        return self.receipt_qty * self.receipt_price

    @classmethod
    def as_form(cls,
                receiptDate: str = Form(...),
                receiptNumber: int = Form(...),
                receiptProduct: str = Form(...),
                receiptQty: int = Form(...),
                receiptPrice: float = Form(...)) -> "ReceiptForm":
        return cls(receiptDate=receiptDate,
                   receiptNumber=receiptNumber,
                   receiptProduct=receiptProduct,
                   receiptQty=receiptQty,
                   receiptPrice=receiptPrice)

    model_config = {
        "populate_by_name": True
    }


@printer_router.post("/billrender", response_class=HTMLResponse)
async def submit_link(request: Request, form=Depends(ReceiptForm.as_form)):
    xlsx_template = BASE_DIR / "api_service/E1.xlsx"
    with open(xlsx_template, "rb") as template_file:
        file_stream = BytesIO(template_file.read())
        writer = BookWriter(file_stream)
    form_dict = form.dict()
    form_dict.update({'total_by_words': num2words(int(form.total), lang='ru')})
    writer.render_book([form_dict])
    result_stream = BytesIO()
    writer.save(result_stream)
    result_stream.seek(0)
    response = StreamingResponse(result_stream,
                                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename=receipt.xlsx"
    return response


@printer_router.post("/get_price_excel")
async def export_data_to_excel(bulk: ParsingResultOut):
    _dt = bulk.dt_parsed.strftime("%d-%m-%Y %H:%M")

    wb = Workbook()
    ws = wb.active
    ws.title = "Price"
    ws.append([f"Актуально: {_dt}"])
    ws.append(["Предложение действительно в течение суток"])
    ws.append([])
    ws.append(["Название", "Гарантия", "Цена"])

    for item in bulk.parsing_result:
        ws.append([
            str(item.title or ""),
            str(item.warranty or ""),
            item.output_price if item.output_price is not None else ""
        ])

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=3, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = column_cells[0].column_letter
        ws.column_dimensions[column_letter].width = max_length + 2

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"by_cifrohub_{bulk.dt_parsed.strftime('%d_%m_%Y')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

    return StreamingResponse(stream,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers=headers)
